from flask import Flask, request
import json
from utils import send_text, send_buttons, send_quick_replies
from responses import RESPONSES
import requests
import os
from openpyxl import load_workbook
from logger import log_message

PAGE_ACCESS_TOKEN = os.getenv("FB_PAGE_TOKEN")
TG_BOT_TOKEN = os.getenv("TG_BOT_TOKEN")
TG_ADMIN_ID = os.getenv("TG_ADMIN_ID")
LOG_FILE = "fb_messages_log.xlsx"
app = Flask(__name__)

def send_telegram_message(text):
    url = f"https://api.telegram.org/bot{TG_BOT_TOKEN}/sendMessage"
    data = {
        "chat_id": TG_ADMIN_ID,
        "text": text,
        "parse_mode": "HTML"
    }
    r = requests.post(url, data=data)
    if r.status_code != 200:
        print(f"[ERROR] Telegram notify failed: {r.text}")
    else:
        print("[SENT] Telegram notification sent")

def get_user_name(user_id):
    url = f"https://graph.facebook.com/{user_id}?fields=first_name,last_name&access_token={PAGE_ACCESS_TOKEN}"
    try:
        r = requests.get(url)
        if r.status_code == 200:
            data = r.json()
            return f"{data.get('first_name', '')} {data.get('last_name', '')}".strip()
        else:
            return "Unknown"
    except Exception:
        return "Unknown"

def is_new_user(user_id: str) -> bool:
    if not os.path.exists(LOG_FILE):
        return True  # файла нет — точно новый юзер

    wb = load_workbook(LOG_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):  # пропускаем заголовки
        logged_id = str(row[1])  # второй столбец — user_id
        if str(user_id) == logged_id:
            return False

    return True

def handle_postback(sender_id, payload):
    name = get_user_name(sender_id)
    log_message(sender_id, name, "", payload)
    response = RESPONSES.get(payload.lower())
    if response:
        send_text(sender_id, response)
    else:
        send_text(sender_id, "Выберите, пожалуйста, кнопку ниже.")
    send_buttons(sender_id)

def handle_message(sender_id, message):
    text = message.get("text", "").lower()
    quick_payload = message.get("quick_reply", {}).get("payload", "").upper()
    name = get_user_name(sender_id)
    log_message(sender_id, name, text, quick_payload)
    if is_new_user(sender_id):
        text = f"📩 <b>Новый пользователь начал диалог</b>:\nИмя: {name}\nID: {sender_id}\nПервое сообщение: {text}"
        send_telegram_message(text)

    if any(greet in text for greet in ["привет", "добрый день", "здравствуйте", "hi", "hello"]):
        send_text(sender_id, RESPONSES["greeting"])
        send_buttons(sender_id)
        send_quick_replies(sender_id)

    elif quick_payload == "CALL_ME":
        send_text(sender_id, "👌 Я уже вас заметил. Мы свяжемся с вами в ближайшее время.")

    elif quick_payload in RESPONSES:
        send_text(sender_id, RESPONSES[quick_payload])
        send_buttons(sender_id)

    else:
        send_text(sender_id, "Не понял вас, выберите из кнопок ниже.")
        send_buttons(sender_id)
        send_quick_replies(sender_id)

@app.route('/webhook', methods=['GET'])
def verify():
    from os import getenv
    VERIFY_TOKEN = getenv("FB_VERIFY_TOKEN", "your_verify_token")
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")
    if mode == "subscribe" and token == VERIFY_TOKEN:
        return challenge, 200
    return "Verification failed", 403

@app.route('/webhook', methods=['POST'])
def webhook():
    data = request.get_json()
    print(f"[LOG] Incoming: {json.dumps(data, indent=2, ensure_ascii=False)}")
    if data.get("object") == "page":
        for entry in data.get("entry", []):
            for event in entry.get("messaging", []):
                sender_id = event["sender"]["id"]
                if event.get("postback"):
                    payload = event["postback"].get("payload")
                    if payload:
                        handle_postback(sender_id, payload)
                elif event.get("message"):
                    handle_message(sender_id, event["message"])
    return "ok", 200

if __name__ == '__main__':
    app.run(debug=True, port=5000)
