from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

LOG_FILE = "fb_messages_log.xlsx"

def log_message(user_id: str, name: str, message: str = "", payload: str = ""):
    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата", "ID пользователя", "Имя", "Сообщение", "Payload"])
        wb.save(LOG_FILE)

    wb = load_workbook(LOG_FILE)
    ws = wb.active

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        user_id,
        name,
        message,
        payload
    ])

    wb.save(LOG_FILE)
